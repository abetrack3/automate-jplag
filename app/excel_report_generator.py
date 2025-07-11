import re
from typing import List, Optional, Tuple

import pandas as pd
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

EXCEL_FILE_NAME: str = 'result.xlsx'


def __extract_student_id_from_file_name(file_name: str) -> Optional[str]:
    pattern: str = r'(\d{4,8})'

    matches = re.findall(pattern, file_name)

    if matches:

        return matches[0]

    else:

        return None


def __custom_sort_key_function(element: Tuple[str, str, str, str, float,]) -> Tuple[str, float, str, str, str,]:
    return (
        element[0] if element[0] is not None else 'z',
        1.0 - (element[4] if element[4] is not None else 0.0),
        element[1] if element[1] is not None else 'z',
        element[2] if element[2] is not None else 'z',
        element[3] if element[3] is not None else 'z',
    )


def __wrap_cells(worksheet: Worksheet) -> None:
    for row in worksheet.iter_rows():

        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


def __write_to_excel_file(data: pd.DataFrame) -> None:
    print('Generating report excel...')

    with pd.ExcelWriter(EXCEL_FILE_NAME) as writer:
        data.to_excel(writer, index=True)

        # adjusting the column widths based on column names
        worksheet: Worksheet = writer.sheets['Sheet1']

        for column_index, column_name in enumerate(data.columns, start=2):
            column_width: int = len(column_name)

            column_letter: str = get_column_letter(column_index)

            worksheet.column_dimensions[column_letter].width = column_width * 1.2

        # Freeze the first row
        worksheet.freeze_panes = 'A2'

        __wrap_cells(worksheet)

    print(F'Generated: "{EXCEL_FILE_NAME}"')


def generate_excel_report() -> None:
    print('\n\nGenerating Excel report from CSV...')

    # Read the CSV file
    df = pd.read_csv('results/results.csv')

    result_set: List[Tuple[str, str, str, str, float]] = []

    for _, row in df.iterrows():
        submission_name_1: str = row['submissionName1']
        submission_name_2: str = row['submissionName2']
        max_similarity: float = row['maxSimilarity']

        if max_similarity == 0.0:
            continue

        student_a_id: str = __extract_student_id_from_file_name(submission_name_1)
        student_b_id: str = __extract_student_id_from_file_name(submission_name_2)

        if student_a_id == student_b_id:
            continue

        result_set.append((
            student_a_id,
            student_b_id,
            submission_name_1,
            submission_name_2,
            round(max_similarity * 100.0, 2),  # Convert to percentage
        ))

    # Sort results using custom sort key (optional — define below if needed)
    result_set.sort(key=__custom_sort_key_function)

    # Unzip result_set into separate lists
    source_student_ids, target_student_ids, source_files, target_files, similarities = zip(*result_set)

    # Create final DataFrame
    data_frame = pd.DataFrame({
        'Student A ID': source_student_ids,
        'Student B ID': target_student_ids,
        'Similarity Percentage': similarities,
        'Student A File Name': source_files,
        'Student B File Name': target_files,
    })

    # Write to Excel
    __write_to_excel_file(data_frame)
